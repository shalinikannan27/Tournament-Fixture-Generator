import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_excel(data):
    """
    data: dict containing current tournament state
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tournament Fixture"

    # ─── STYLES ───────────────────────────────────────────────────────────────
    GREEN_HEADER = '1A3A2A'
    WHITE = 'FFFFFF'
    BORDER_COLOR = '000000'
    
    title_font = Font(name='Segoe UI', size=14, bold=True)
    header_font = Font(name='Segoe UI', size=11, bold=True, color=WHITE)
    normal_font = Font(name='Segoe UI', size=10)
    
    thin = Side(border_style='thin', color=BORDER_COLOR)
    medium = Side(border_style='medium', color=BORDER_COLOR)
    dotted = Side(border_style='dotted', color='999999') # Guide for manual writing
    
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    match_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', indent=1)

    # ─── PARAMETERS ──────────────────────────────────────────────────────────
    t_name = data.get('tname', 'Tournament')
    sport = data.get('sport', '') # Category title
    dates = data.get('tdates', '')
    venue = data.get('tvenue', '')
    players = data.get('players', [])
    round_names = data.get('round_names', [])
    num_rows = len(players)
    
    # ─── COLUMN WIDTHS ────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    for i in range(len(round_names)):
        col_letter = chr(ord('C') + i)
        ws.column_dimensions[col_letter].width = 15

    # ─── HEADER ROWS ──────────────────────────────────────────────────────────
    last_col_idx = 2 + len(round_names)
    last_col_letter = chr(ord('A') + last_col_idx - 1)
    
    # Tournament Title
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'].value = t_name.upper()
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Metadata
    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'].value = f"{dates} | {venue}"
    ws['A2'].alignment = center_align
    
    # Category (e.g. Mens Singles Draw)
    ws.merge_cells(f'A3:{last_col_letter}3')
    ws['A3'].value = sport.upper()
    ws['A3'].font = Font(size=12, bold=True, underline='single')
    ws['A3'].alignment = center_align

    # Column Headers
    header_row = 4
    headers = ["SL NO", "NAME OF THE PLAYER"] + round_names
    for c_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c_idx)
        cell.value = h_text
        cell.font = header_font
        cell.fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border

    # ─── BODY DATA ────────────────────────────────────────────────────────────
    start_row = 5
    # First, fill the base player data columns (SL, Name)
    for r_idx, p in enumerate(players):
        row_num = start_row + r_idx
        
        # SL
        c = ws.cell(row=row_num, column=1, value=p.get('sl'))
        c.alignment = center_align
        c.border = thin_border
        
        # Name
        c = ws.cell(row=row_num, column=2, value=p.get('name'))
        c.alignment = left_align
        c.border = thin_border

    # Now handle the Merged Round Columns (The Tournament Tree)
    for c_idx, nm in enumerate(round_names):
        col_idx = 3 + c_idx
        row_span = 2 ** (c_idx + 1)
        
        for r_idx in range(0, num_rows, row_span):
            s_row = start_row + r_idx
            e_row = s_row + row_span - 1
            
            # Merge cell
            if s_row != e_row:
                ws.merge_cells(start_row=s_row, start_column=col_idx, end_row=e_row, end_column=col_idx)
            
            # Style the merged cell
            top_cell = ws.cell(row=s_row, column=col_idx)
            
            # Add a dotted guide line for manual writing
            top_cell.value = "................" 
            top_cell.alignment = Alignment(horizontal='center', vertical='center')
            top_cell.font = Font(color='BBBBBB')
            
            # Apply border to the entire merged range block
            for r in range(s_row, e_row + 1):
                ws.cell(row=r, column=col_idx).border = thin_border

    # Final perimeter border (Medium)
    # Applying medium border to the bounding box from Row 4 to last player row
    for r in range(4, start_row + num_rows):
        ws.cell(row=r, column=1).border = Border(left=medium, right=thin, top=ws.cell(row=r, column=1).border.top, bottom=ws.cell(row=r, column=1).border.bottom)
        ws.cell(row=r, column=last_col_idx).border = Border(right=medium, left=thin, top=ws.cell(row=r, column=last_col_idx).border.top, bottom=ws.cell(row=r, column=last_col_idx).border.bottom)
    
    # Top and bottom medium edges
    for c in range(1, last_col_idx + 1):
        ws.cell(row=4, column=c).border = Border(top=medium, left=ws.cell(row=4, column=c).border.left, right=ws.cell(row=4, column=c).border.right, bottom=ws.cell(row=4, column=c).border.bottom)
        ws.cell(row=start_row + num_rows - 1, column=c).border = Border(bottom=medium, left=ws.cell(row=start_row+num_rows-1, column=c).border.left, right=ws.cell(row=start_row+num_rows-1, column=c).border.right, top=ws.cell(row=start_row+num_rows-1, column=c).border.top)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
